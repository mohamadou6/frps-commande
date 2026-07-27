import unicodedata
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.shortcuts import render
from openpyxl import load_workbook

from accounts.decorators import admin_frps_required, formation_sanitaire_required

from .models import Magasin, Produit


@formation_sanitaire_required
def liste(request):
    query = request.GET.get("q", "").strip()
    produits = Produit.objects.filter(
        actif=True, magasin__in=[Magasin.PRINCIPAL, Magasin.UCPC], stock_disponible__gt=0
    ).order_by("nom")
    if query:
        produits = produits.filter(nom__icontains=query)
    return render(request, "catalogue/liste.html", {"produits": produits, "query": query})


def _normaliser_entete(valeur):
    if valeur is None:
        return ""
    texte = unicodedata.normalize("NFD", str(valeur).strip().lower())
    return "".join(c for c in texte if not unicodedata.combining(c))


def _decimal(valeur):
    if valeur is None or valeur == "":
        return Decimal("0.00")
    try:
        return Decimal(str(valeur).strip()).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


def _lire_fichier_produits(fichier):
    """Lit un classeur Excel dont l'en-tête contient une colonne code produit
    (ex. code_sage), une colonne prix (ex. prix_unitaire) et une colonne stock
    (ex. stock_disponible), dans n'importe quel ordre. Retourne un dict
    {code_sage: (prix, stock)}. Lève ValueError si l'en-tête attendu est introuvable."""
    classeur = load_workbook(filename=fichier, data_only=True, read_only=True)
    feuille = classeur.active
    lignes = feuille.iter_rows(values_only=True)

    entetes = next(lignes, None)
    if entetes is None:
        raise ValueError("Le fichier est vide.")

    colonnes = {}
    for index, entete in enumerate(entetes):
        cle = _normaliser_entete(entete)
        if "code" in cle and "code_sage" not in colonnes:
            colonnes["code_sage"] = index
        elif "prix" in cle and "prix_unitaire" not in colonnes:
            colonnes["prix_unitaire"] = index
        elif "stock" in cle and "stock_disponible" not in colonnes:
            colonnes["stock_disponible"] = index

    manquantes = [c for c in ("code_sage", "prix_unitaire", "stock_disponible") if c not in colonnes]
    if manquantes:
        raise ValueError(
            "Colonnes introuvables dans l'en-tête du fichier : "
            + ", ".join(manquantes)
            + ". Le fichier doit avoir une colonne code produit (ex. code_sage), "
            "une colonne prix (ex. prix_unitaire) et une colonne stock "
            "(ex. stock_disponible)."
        )

    updates = {}
    for ligne in lignes:
        if ligne is None or all(valeur is None for valeur in ligne):
            continue
        code_sage = ligne[colonnes["code_sage"]]
        if code_sage is None or not str(code_sage).strip():
            continue
        code_sage = str(code_sage).strip()
        prix = _decimal(ligne[colonnes["prix_unitaire"]])
        stock = int(_decimal(ligne[colonnes["stock_disponible"]]))
        updates[code_sage] = (prix, stock)
    return updates


@admin_frps_required
def importer_produits(request):
    resultats = None

    if request.method == "POST":
        fichier = request.FILES.get("fichier")
        if not fichier:
            messages.error(request, "Aucun fichier sélectionné.")
        elif not fichier.name.lower().endswith((".xlsx", ".xlsm")):
            messages.error(request, "Le fichier doit être un classeur Excel (.xlsx).")
        else:
            try:
                updates = _lire_fichier_produits(fichier)
            except ValueError as exc:
                messages.error(request, str(exc))
            else:
                produits = {p.code_sage: p for p in Produit.objects.filter(code_sage__in=updates.keys())}

                modifies = []
                introuvables = []
                nb_inchanges = 0
                a_sauvegarder = []

                for code_sage, (prix, stock) in updates.items():
                    produit = produits.get(code_sage)
                    if produit is None:
                        introuvables.append(code_sage)
                        continue
                    if produit.prix_unitaire == prix and produit.stock_disponible == stock:
                        nb_inchanges += 1
                        continue
                    modifies.append(
                        {
                            "code_sage": code_sage,
                            "nom": produit.nom,
                            "ancien_prix": produit.prix_unitaire,
                            "nouveau_prix": prix,
                            "ancien_stock": produit.stock_disponible,
                            "nouveau_stock": stock,
                        }
                    )
                    produit.prix_unitaire = prix
                    produit.stock_disponible = stock
                    a_sauvegarder.append(produit)

                if a_sauvegarder:
                    Produit.objects.bulk_update(a_sauvegarder, ["prix_unitaire", "stock_disponible"], batch_size=200)

                resultats = {
                    "total_lignes": len(updates),
                    "modifies": modifies,
                    "nb_modifies": len(modifies),
                    "nb_inchanges": nb_inchanges,
                    "introuvables": introuvables,
                    "nb_introuvables": len(introuvables),
                }

                if modifies:
                    messages.success(request, f"{len(modifies)} produits mis à jour.")
                else:
                    messages.info(request, "Aucun changement détecté dans ce fichier.")

    return render(request, "catalogue/importer.html", {"resultats": resultats})
